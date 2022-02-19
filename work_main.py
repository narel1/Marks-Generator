# Name: Anushkha Singh
# Roll No.: 1901ME72

# Filename: work_main.py : Contains functionalities of the web application for error handling, generating marksheets and sending email.

# Import relevant libraries
import os
import json
import csv
import openpyxl
import shutil
import copy
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Function for handling file errors
def error_handling(file, upload_folder):
	try:
		if file.filename !='':
			file.save(os.path.join(upload_folder, file.filename))

	except:
		return r"Please close the Excel files."

# Function to generate marksheet for all roll numbers present in masters_roll csv file
def generate_marksheet(path1, path2, pos, neg):

	# Get the data fetched from user written in params.json file
	pos = float(pos)
	neg = float(neg)

	dict_input = {
	               "path1": path1,
	               "path2": path2,
	               "positive": pos,
	               "negative": neg,
	             }

    # Write the input parameters to the file
	with open("params.json", "w") as f:
		json.dump(dict_input, f)

	# Create the necessary directories in case they don't exist
	if not os.path.isdir("sample_output"):
		os.mkdir("sample_output")

	if not os.path.isdir(os.path.join("sample_output", "marksheet")):
		os.mkdir(os.path.join("sample_output", "marksheet"))

	if path1.split('\\')[-1] != "master_roll.csv":
		if os.path.isfile(path1):
			os.remove(path1)

	# Create the roll_name_map dictionary from masters_roll csv whose keys are the roll nos and values are the names.
	with open(path1, "r") as csvfile:

		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		roll_name_map= {}

		for row in reader:
			if row[0] != "roll":
				roll_name_map[row[0]] = row[1]

	# Create the stud_info dictionary from responses csv whose keys are the roll nos and values are the info of the students.
	with open(path2, "r") as csvfile:

		stud_info = {}
		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		header = next(reader)

		f_header = []
		f_header.extend(header[1:5])
		f_header.append(header[6])
		f_header.append("Options")

		for row in reader:

			stud_info[row[6]] = {}
			i=1

			for h in f_header:

				if i==4:
					break
				stud_info[row[6]][h] = row[i]
				i+=1

			stud_info[row[6]][f_header[4]] = row[6]
			stud_info[row[6]][f_header[5]] = row[7:]

	# Make sure whether "ANSWER" keyword is present in responses csv file
	if "ANSWER" not in stud_info.keys():

		if path2.split('\\')[-1] != "responses.csv":

			if os.path.isfile(path2):
				os.remove(path2)

		return r"no roll number with ANSWER is present, Cannot Process!"

	correct_answers = stud_info["ANSWER"]["Options"]
	# print(correct_answers)
	total_marks = len(correct_answers)*pos

	# Include the roll numbers that are present in masters_roll csv but not in responses csv in stud_info dictionary
	for roll in roll_name_map.keys():

		if roll not in stud_info.keys():

			stud_info[roll] = {}
			stud_info[roll]["Name"] = roll_name_map[roll]
			stud_info[roll]["Roll Number"] = roll
			stud_info[roll]["Options"] = ['']*len(correct_answers)

	# Create f_stud_data which contains the final info for all the students like marks before negative, marks after 
	# negative, correct answers, wrong answers, missing answers etc
	f_stud_data = {}
	for roll in stud_info.keys():

		if len(stud_info[roll]["Options"]) > 50:
			return r"This application will accept MAX 50 questions."

		correct = 0
		missing = 0
		incorrect = 0

		f_stud_data[roll] = {}
		f_stud_data[roll]["Name"] = stud_info[roll]["Name"]
		f_stud_data[roll]["Roll Number"] = stud_info[roll]["Roll Number"]

		for idx,val in enumerate(stud_info[roll]["Options"]):

			if val == correct_answers[idx]:
				correct += 1

			elif val != correct_answers[idx]:

				if val == "":
					missing += 1
				else:
					incorrect += 1

		marks_stud = (correct* pos) + (incorrect* neg)

		f_stud_data[roll]["right_no"] = correct
		f_stud_data[roll]["wrong_no"] = incorrect
		f_stud_data[roll]["not_attempt"] = missing
		f_stud_data[roll]["max_no"] = correct + incorrect + missing
		f_stud_data[roll]["total_right"] = correct* pos
		f_stud_data[roll ]["total_wrong"] = incorrect* neg
		f_stud_data[roll]["total_max"] = str(marks_stud) + "/" + str(total_marks)

	# Iterate through all the keys of f_stud_data dictionary and create excel marksheets for all the roll numbers
	for roll in f_stud_data.keys():

		# Make the roll number in upper case, eg 1901c10->>> 1901CS10 
		roll_f =""

		for x in roll:

			if not (ord(x) >=48 and ord(x)<=57):
				roll_f+= x.upper()
			else:
				roll_f+=x

		# Initialize a new excel workbook
		wb = Workbook()
		sheet = wb.active

		row_count = 60
		column_count = 5

		# Write data in the excel sheet
		for i in range(1, row_count + 1):

			for j in range(1, column_count + 1):

				sheet.column_dimensions[get_column_letter(j)].width = 17
				sheet.cell(row=i, column=j).font = Font(name= "Century", size=12)

		sheet.title = "quiz"

		# Write and format the data in the sheet
		img = openpyxl.drawing.image.Image('IITP_Logo.jpg')
		img.anchor = 'A1'
		sheet.add_image(img)
		sheet.merge_cells('A5:E5')

		sheet["A5"].value = 'Mark Sheet'
		sheet["A5"].font = Font(name="Century",size=18, bold=True, underline = "single")
		sheet["A5"].border = Border()
		sheet["A5"].alignment = Alignment(horizontal="center", vertical="bottom")

		sheet["A6"].value = "Name:"
		sheet["A6"].alignment = Alignment(horizontal="right")
		sheet["B6"].value = f_stud_data[roll]["Name"]
		sheet["B6"].font = Font(name= "Century", size=12, bold=True)

		sheet["B6"].alignment = Alignment(horizontal= "left")
		sheet["D6"].value = "Exam:"
		sheet["D6"].alignment = Alignment(horizontal="right")
		sheet["E6"].value = "quiz"

		sheet["E6"].font = Font(name= "Century", size=12, bold=True)
		sheet["E6"].alignment = Alignment(horizontal= "left")
		sheet["A7"].value = "Roll Number:"
		sheet["A7"].alignment = Alignment(horizontal="right")

		sheet["B7"].value = f_stud_data[roll]["Roll Number"]
		sheet["B7"].font = Font(name= "Century", size=12, bold=True)
		sheet["B7"].alignment = Alignment(horizontal= "left")

		sheet.append(["","","","",""])
		marks_info = []

		marks_info.append(tuple(["", "Right", "Wrong", "Not Attempt", "Max"]))
		marks_info.append(tuple(["No.", f_stud_data[roll]["right_no"], f_stud_data[roll]["wrong_no"],
		                    f_stud_data[roll]["not_attempt"], f_stud_data[roll]["max_no"]]))

		marks_info.append(tuple(["Marking", pos, neg, 0, ""]))
		marks_info.append(tuple(["Total", f_stud_data[roll]["total_right"], f_stud_data[roll]["total_wrong"],
			               "", f_stud_data[roll]["total_max"]]))

		marks_info = tuple(marks_info)
		thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

		i1=0

		for i in range(9, 13):

			j1=0

			for j in range(1, 6):

				sheet.cell(row=i, column=j).value = marks_info[i1][j1]
				sheet.cell(row=i, column=j).border = thin_border
				sheet.cell(row=i, column=j).alignment = Alignment(horizontal= "center")

				if i1==0 or (j1 == 0 and i1>0):
					sheet.cell(row=i, column=j).font = Font(bold = True, name= "Century", size=12)

				if i1 > 0:

					if j1 == 1:
						sheet.cell(row=i, column=j).font = Font(color = "00008000", name= "Century", size=12)
					elif j1 == 2:
						sheet.cell(row=i, column=j).font = Font(color = "00FF0000", name= "Century", size=12)

				if i1 == 3 and j1 == 4:
					sheet.cell(row=i, column=j).font = Font(color = "000000FF", name= "Century", size=12)

				j1+=1

			i1+=1

		sheet.append(["","","","",""])
		sheet.append(["","","","",""])

		s1 = copy.deepcopy(stud_info)
		s_ans = s1[roll]["Options"]
		c_ans = s1["ANSWER"]["Options"]
		c1 = copy.deepcopy(c_ans)

		# Mention the students marked options and the correct answers in the excel worbook
		turn =1

		while 1:
			
			if len(s_ans) == 0:
				break

			stud_ans = []
			stud_ans.append(tuple(["Student Ans", "Correct Ans"]))

			if len(s_ans) < 25:
				f_s_ans = s_ans[:len(s_ans)]

			elif len(s_ans) >= 25:
				f_s_ans = s_ans[:25]

			if len(c1) < 25:
				f_c_ans = c1[:len(c1)]

			elif len(c1) >= 25:
				f_c_ans = c1[:25]

			for k in range(len(f_s_ans)):
				s_ans.pop(0)

			for k1 in range(len(f_c_ans)):
				c1.pop(0)

			for x in range(len(f_c_ans)):
				stud_ans.append(tuple([f_s_ans[x], f_c_ans[x]]))

			stud_ans = tuple(stud_ans)
			end = 15+ len(stud_ans)

			thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

			i1=0

			for i in range(15, end):

				j1=0

				for j in range(turn, turn+2):

					sheet.cell(row=i, column=j).value =stud_ans[i1][j1]
					sheet.cell(row=i, column=j).border = thin_border
					sheet.cell(row=i, column=j).alignment = Alignment(horizontal= "center")

					if i1==0:
						sheet.cell(row=i, column=j).font = Font(bold = True, name= "Century", size=12)

					if i1>0:

						if (stud_ans[i1][0] == stud_ans[i1][1]):
							sheet.cell(row=i, column=j).font = Font(color = "00008000", name= "Century", size=12)
						else:
							if stud_ans[i1][0] != '':
								sheet.cell(row=i, column=j).font = Font(color = "00FF0000", name= "Century", size=12)

					if i1>0 and j1==1:
						sheet.cell(row=i, column=j).font = Font(color = "000000FF", name= "Century", size=12)

					j1+=1

				i1+=1

			turn+=3

		# Save the workbook to the marksheet folder within the sample output folder
		name = os.path.join(os.getcwd(), "sample_output", "marksheet",(roll_f+".xlsx"))

        # File error handling
		try:
			wb.save(name)
		except:
			return r"Please close the Excel files."

# Generate a concise marksheet containing the info of all the students
def concise_marksheet(path1, path2, pos, neg):

	# Get the data fetched from user written in params.json file
	pos = float(pos)
	neg = float(neg)

	dict_input = {
	               "path1": path1,
	               "path2": path2,
	               "positive": pos,
	               "negative": neg,
	             }

    # Write the input parameters in a file
	with open("params.json", "w") as f:
		json.dump(dict_input, f)

	if path1.split('\\')[-1] != "master_roll.csv":
		if os.path.isfile(path1):
			os.remove(path1)

	# Create a roll_name_map dictionary from masters_roll csv whose keys are the roll nos and values are the names.
	with open(path1, "r") as csvfile:

		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		roll_name_map= {}

		for row in reader:
			if row[0] != "roll":
				roll_name_map[row[0]] = row[1]

	# Create a stud_info dictionary from responses csv whose keys are the roll nos and values are the info of the students.
	with open(path2, "r") as csvfile:

		stud_info = {}
		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		header = next(reader)

		f_header = []
		f_header.extend(header[0:7])
		f_header.append("Options")

		for row in reader:

			stud_info[row[6]] = {}
			i=0

			for h in f_header:

				if i==7:
					break

				stud_info[row[6]][h] = row[i]
				i+=1

			stud_info[row[6]][f_header[7]] = row[7:]

	# Make sure whether "ANSWER" keyword is present in responses csv file
	if "ANSWER" not in stud_info.keys():

		if path2.split('\\')[-1] != "responses.csv":

			if os.path.isfile(path2):
				os.remove(path2)

		return r"no roll number with ANSWER is present, Cannot Process!"

	correct_answers = stud_info["ANSWER"]["Options"]
	total_marks = len(correct_answers)*pos

	# Create f_stud_data which contains the final info for all the students like marks before negative, marks after 
	# negative, correct answers, wrong answers, missing answers etc
	f_stud_data = {}

	for roll in stud_info.keys():

		if len(stud_info[roll]["Options"]) > 50:
			return r"This application will accept MAX 50 questions."

		correct = 0
		missing = 0
		incorrect = 0

		f_stud_data[roll] = {}
		f_stud_data[roll]["Name"] = stud_info[roll]["Name"]
		f_stud_data[roll]["Roll Number"] = stud_info[roll]["Roll Number"]

		for idx,val in enumerate(stud_info[roll]["Options"]):

			if val == correct_answers[idx]:
				correct += 1

			elif val != correct_answers[idx]:

				if val == "":
					missing += 1
				else:
					incorrect += 1

		marks_stud = (correct* pos) + (incorrect* neg)
		f_stud_data[roll]["status_Ans"] = str([correct, incorrect, missing])
		f_stud_data[roll]["score"] = str((correct* pos)) + "/" + str(total_marks)
		f_stud_data[roll]["score_after_neg"] = str(marks_stud) + "/" + str(total_marks)

	concise_info = []
	header =["Timestamp", "Email address", "Google Score", "Name", "IITP webmail", "Phone (10 digit only)", "Score_After_Negative", "Roll Number"]
	end = 7+ (len(correct_answers)-1)

	for i in range(7, end+1):
		header.append("Unnamed: "+ str(i))
	header.append("statusAns")

	# Iterate through all the roll numbers in stud_info dictionary and append the data of all the roll numbers in
	# concise_info list
	for roll in stud_info.keys():

		s_inf = []
		s_inf.append(stud_info[roll]["Timestamp"])
		s_inf.append(stud_info[roll]["Email address"])
		s_inf.append(f_stud_data[roll]["score"])
		s_inf.append(stud_info[roll]["Name"])
		s_inf.append(stud_info[roll]["IITP webmail"])

		s_inf.append(stud_info[roll]["Phone (10 digit only)"])
		s_inf.append(f_stud_data[roll]["score_after_neg"])
		s_inf.append(stud_info[roll]["Roll Number"])
		s_inf.extend(stud_info[roll]["Options"])
		s_inf.append(f_stud_data[roll]["status_Ans"])
		concise_info.append(s_inf)

	# Append the data of absent students (students present in masters_roll csv file but not in responses csv 
	# file) to concise_info list which is to be entered in the concise__marksheet csv file
	for roll in roll_name_map.keys():

		if roll not in stud_info.keys():

			x = ["","","Absent", roll_name_map[roll], "","", "Absent", roll]
			x.extend([""]*len(correct_answers))
			concise_info.append(x)

	# Create the necessary directories in case they don't exist
	if not os.path.isdir("sample_output"):
		os.mkdir("sample_output")

	if not os.path.isdir("sample_output//marksheet"):
		os.mkdir("sample_output//marksheet")

	# Create the concise_marksheet csv file by opening in write mode and writing header and rows to the file
	# from the concise_info list
	try:
		with open(os.path.join("sample_output", "marksheet","concise_marksheet.csv"), "w", newline='') as f:
			csvwriter = csv.writer(f)
			csvwriter.writerow(header)
			csvwriter.writerows(concise_info)
	except:
			return r"Please close the Excel files."

# Email function to send the marks along with body, subject, attachment from sender to receiver
def email_func(sender, receiver, filename, path, pos, neg):

	# Instance of MIMEMultipart
	msg = MIMEMultipart()

	# Store the senders email address  
	msg['From'] = sender

	# Store the receivers email address 
	msg['To'] = receiver

	# Store the subject 
	msg['Subject'] = "Quiz Marksheet"

	# String to store the body of the mail
	body = f" \
	Dear Student,\n\n \
	CS384 2021 Quiz marks are attached for reference.\n \
	{pos} for Correct, {neg} for wrong.\n \
	-- \n \
	Anushkha Singh \
                       \
	"

	# Attach the body with the msg instance
	msg.attach(MIMEText(body, 'plain'))

	# Open the file to be sent 
	filename = filename
	attachment = open(path, "rb")

	# Instance of MIMEBase and named as p
	p = MIMEBase('application', 'octet-stream')

	# Change the payload into encoded form
	p.set_payload((attachment).read())

	# Encode into base64
	encoders.encode_base64(p)

	p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

	# Attach the instance 'p' to 'msg' instance
	msg.attach(p)

	# Create SMTP session
	s = smtplib.SMTP('stud.iitp.ac.in', 587)

	# Start TLS for security
	s.ehlo()
	s.starttls()
	s.ehlo

	# Authentication for logging into the server
	s.login(sender, "abc")          # PLEASE ENTER YOUR IITP EMAIL PASSWORD INSTEAD OF abc

	# Convert the Multipart msg into a string
	text = msg.as_string()

	# Send the email
	s.sendmail(sender, receiver, text)

	# Terminate the session
	s.close()

# Main function to send email with file to all students who were present (students present in responses csv file)
def Send_email():

	f = open ('params.json', "r")

	# Read from file
	data = json.loads(f.read())

	# Get the data entered by the user into the form from params.json file
	path1 = data["path1"]
	path2 = data["path2"]
	pos = data["positive"]
	neg = data["negative"]

	# Check the necessary conditons whether the sample_output/matksheet directory exists or any files within the 
	# marksheet folder exists 
	if os.path.isdir("sample_output"):

		if os.path.isdir("sample_output//marksheet"):

			if ((len(os.listdir("sample_output//marksheet")) == 0)):
				return r"There is nothing to send !! Please generate marksheet first"

		else:
			return r"There is nothing to send !! Please generate marksheet first"

	else:
		return r"There is nothing to send !! Please generate marksheet first"

	if path1.split('\\')[-1] != "master_roll.csv":
		if os.path.isfile(path1):
			os.remove(path1)

	# Create a roll_name_map dictionary from masters_roll csv whose keys are the roll nos and values are the names
	with open(path1, "r") as csvfile:

		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		roll_name_map= {}

		for row in reader:

			if row[0] != "roll":
				roll_name_map[row[0]] = row[1]

	# Create stud_info dictionary from responses csv whose keys are the roll nos and values are the info of the students
	with open(path2, "r") as csvfile:

		stud_info = {}
		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		header = next(reader)

		f_header = []
		f_header.extend(header[0:7])
		f_header.append("Options")

		for row in reader:

			stud_info[row[6]] = {}
			i=0

			for h in f_header:

				if i==7:
					break

				stud_info[row[6]][h] = row[i]
				i+=1

			stud_info[row[6]][f_header[7]] = row[7:]

	# Make sure whether "ANSWER" keyword is present in responses csv file
	if "ANSWER" not in stud_info.keys():

		if path2.split('\\')[-1] != "responses.csv":

			if os.path.isfile(path2):
				os.remove(path2)

		return r"no roll number with ANSWER is present, Cannot Process!"
	
	# Iterate through the keys which are the roll nos of stud_info dictionary and send email to all the roll numbers
	# whose file exists. Emails are sent to both IITP webmail and student's email id.
	try:

		for roll in stud_info.keys():

			if roll != "ANSWER":

				filename = os.path.join("sample_output", "marksheet", (roll+ ".xlsx"))

				if os.path.isfile(filename):

					# PLEASE ENTER YOUR IITP EMAIL ID INSTEAD OF xyz
					email_func("xyz", stud_info[roll]["Email address"], roll+ ".xlsx", filename, pos, neg)
					email_func("xyz", stud_info[roll]["IITP webmail"], roll+ ".xlsx", filename,  pos, neg)

	except:
			return r"Please enter valid login credentials!"